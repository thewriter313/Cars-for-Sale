from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import openpyxl
import pandas as pd
import os.path

class Car:
    
    def __init__(self, make):
        self.make = make

    def export(self, dataframe):
        df = pd.DataFrame(data=dataframe)
        new_header = df.iloc[0]
        df = df[1:]
        df.columns = new_header

        if os.path.exists(str(self.make) + '.xlsx'):
            print('Exporting ' + self.make + ' car\'s to Excel file: \'' + str(self.make) + '.xlsx\'')
            with pd.ExcelWriter(str(self.make) + '.xlsx', engine='openpyxl', mode='a', if_sheet_exists = 'replace') as writer:
                df.to_excel(writer, sheet_name=self.make)
                
        else:
            print('Creating Excel file: \'' + self.make + '.xlsx\'')
            wb = Workbook()
            ws = wb.active
            ws.title = self.make
            wb.save(str(self.make) + '.xlsx')

            print('Exporting ' + self.make + ' car\'s to Excel file: \'' + str(self.make) + '.xlsx\'')
            with pd.ExcelWriter(str(self.make) + '.xlsx', engine='openpyxl', mode='a', if_sheet_exists = 'replace') as writer:
                df.to_excel(writer, sheet_name=self.make)
        print('Data Saved in file ' + str(self.make) + '.xlsx')

        workbook = openpyxl.load_workbook(str(self.make) + '.xlsx')
        worksheet = workbook.active
        worksheet['A1'] = 'No.'

        max_row = worksheet.max_row

        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 10
        worksheet.column_dimensions['C'].width = 15        
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 10
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 100

        for i in range(2, max_row+1):
            for cell in worksheet[str(i) + ':' + str(i)]:
                cell.font = Font(bold=False, size=12)
                cell.border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

        for cell in worksheet['1:1']:
            cell.font = Font(bold=True, size=14)

        workbook.save(str(self.make) + '.xlsx')
        
        return df
        